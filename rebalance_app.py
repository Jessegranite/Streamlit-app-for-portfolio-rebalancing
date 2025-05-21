
import pandas as pd
import os
import glob
import streamlit as st
import matplotlib.pyplot as plt
import openpyxl

# === CONFIG ===
st.set_page_config(page_title="Portfolio Rebalancer", layout="wide")

st.markdown("""
    <style>
        body {
            background-color: #f4f1ee;
            color: #3e3c3d;
            font-family: 'Inter', sans-serif;
        }
        .stButton>button, .stDownloadButton>button {
            background-color: #ada69e;
            color: white;
            border-radius: 6px;
            font-weight: 500;
        }
        .stDataFrame thead {
            background-color: #dcd6cf;
            color: black;
        }
        .block-label {
            font-weight: 600;
            font-size: 1rem;
            padding: 0.25rem 0;
            text-align: center;
        }
        .st-expanderContent {
        background-color: #f8f5f0 !important;
    }
</style>
""", unsafe_allow_html=True)

st.title("📊 Pierre's Portfolio Rebalancer")

# === File uploader ===
uploaded_file = st.file_uploader("Upload an Excel file", type=[".xlsx"])
if uploaded_file is None:
    st.stop()

# === Extract client name from A3 ===
wb = openpyxl.load_workbook(uploaded_file)
ws = wb.active
client_name = ws["A3"].value or "Client"

ws = wb.active
client_name = ws["A3"].value or "Client"
st.markdown(f"""
    <div style='background-color: #eae6e1; padding: 1rem; border-radius: 10px; text-align: center;
    font-size: 1.6rem; font-weight: bold; color: #3e3c3d; border: 1px solid #ccc; margin-top: 1rem; margin-bottom: 0.5rem;'>
        👤 {client_name}
    </div>
""", unsafe_allow_html=True)



# === Load and clean data ===
df = pd.read_excel(uploaded_file)
df.drop(df.columns[0], axis=1, inplace=True)
df.rename(columns={
    df.columns[0]: "Asset Class",
    df.columns[1]: "Security Name",
    df.columns[2]: "Quantity",
    df.columns[3]: "Market Value (CAD)"
}, inplace=True)

df["Asset Class"] = df["Asset Class"].ffill().str.strip()
df = df[df["Quantity"].notna()]

# === Define desired order ===
desired_order = ["Cash & Cash Equivalents", "Bonds", "Canadian Equity", "Global Equity"]

# === Compute asset-class-level stats ===
grouped = df.groupby("Asset Class")["Market Value (CAD)"].sum().reset_index()
grouped.columns = ["Asset Class", "Current $"]
total_value = grouped["Current $"].sum()
st.markdown(f"<div style='text-align:center; font-size:1.1rem; color:#5a5959; margin-bottom:1.5rem;'>Total Portfolio Value: <strong>${total_value:,.2f}</strong></div>", unsafe_allow_html=True)
grouped = df.groupby("Asset Class")["Market Value (CAD)"].sum().reset_index()
grouped.columns = ["Asset Class", "Current $"]
total_value = grouped["Current $"].sum()
grouped["Current %"] = grouped["Current $"] / total_value * 100

# === Security-level grouping ===
security_df = df.groupby(["Asset Class", "Security Name"])["Market Value (CAD)"].sum().reset_index()
security_df["Class Total"] = security_df.groupby("Asset Class")["Market Value (CAD)"].transform("sum")
security_df["Current %"] = security_df["Market Value (CAD)"] / security_df["Class Total"]


# === Asset Class Input UI ===
sec_methods, sec_inputs, sec_locks = {}, {}, {}
st.subheader("🎯 Asset Class Allocation Targets")

# Line under the section title
st.markdown("<hr style='border: 1px solid #000;'>", unsafe_allow_html=True)

methods, inputs, locks = {}, {}, {}

for asset in desired_order:
    if asset not in grouped["Asset Class"].values:
        continue

    st.markdown(f"**{asset}**", unsafe_allow_html=True)
    row = grouped[grouped['Asset Class'] == asset]
    current_val = row["Current $"].values[0] if not row.empty else 0

    cols = st.columns([2.5, 3.5, 1.2, 1])
    with cols[0]:
        methods[asset] = st.selectbox("Method", ["%", "$", "$ ∆"], key=f"method_{asset}")
    with cols[1]:
        if methods[asset] == "%":
            default_val = row["Current %"].values[0] if not row.empty else 0
        elif methods[asset] == "$":
            default_val = current_val
        else:
            default_val = 0.0
        inputs[asset] = st.number_input("Target", value=default_val, step=100.0 if methods[asset] != "%" else 0.1, key=f"val_{asset}")
    with cols[2]:
        locks[asset] = st.toggle("Lock", value=False, key=f"lock_{asset}")

    asset_securities = security_df[security_df["Asset Class"] == asset]
    with st.expander(f"🔽 Set Targets for Securities in {asset}"):
        st.markdown("""
            <style>
                [data-testid='stExpander'] > div > div {
                    background-color: #f8f5f0;
                    padding: 1rem;
                    border-radius: 10px;
                }
            </style>
        """, unsafe_allow_html=True)
        for _, row in asset_securities.iterrows():
            sec = row["Security Name"]
            key = f"{asset}_{sec}"
            cols = st.columns([3, 1.5, 2.5, 1])
            with cols[0]:
                st.markdown(f"{sec} @ {row['Current %']:.2%}")
            with cols[1]:
                sec_methods[key] = st.selectbox("", ["%", "$", "$ ∆"], key=f"smethod_{key}")
            with cols[2]:
                method_init = sec_methods.get(key, "%")
                current_val = row["Market Value (CAD)"]
                if method_init == "%":
                    default_val = row["Current %"] * 100
                elif method_init == "$":
                    default_val = current_val
                else:
                    default_val = 0.0
                sec_inputs[key] = st.number_input("", value=default_val, step=100.0 if method_init != "%" else 0.1, key=f"sval_{key}")
            with cols[3]:
                sec_locks[key] = st.toggle("Lock", value=False, key=f"slock_{key}")

    # Line after each asset class block
    st.markdown("<hr style='border: 1px solid #000;'>", unsafe_allow_html=True)

# === Rebalancing logic ===
target_dollars, locked_assets, unlocked_assets = {}, [], []
for _, row in grouped.iterrows():
    asset = row["Asset Class"]
    current = row["Current $"]
    method = methods[asset]
    val = inputs[asset]
    if locks[asset]:
        if method == "%":
            target_dollars[asset] = val / 100 * total_value
        elif method == "$":
            target_dollars[asset] = val
        elif method == "$ Δ":
            target_dollars[asset] = current + val
        locked_assets.append(asset)
    else:
        unlocked_assets.append(asset)

assigned = sum(target_dollars.values())
remaining = total_value - assigned
if unlocked_assets:
    even = remaining / len(unlocked_assets)
    for asset in unlocked_assets:
        target_dollars[asset] = even

grouped["Target $"] = grouped["Asset Class"].map(target_dollars)
grouped["Target %"] = grouped["Target $"] / total_value * 100
grouped["Buy/Sell $"] = grouped["Target $"] - grouped["Current $"]

# === Compute security-level targets ===
results = []

for asset in desired_order:
    asset_securities = security_df[security_df["Asset Class"] == asset].copy()
    asset_target = target_dollars.get(asset, 0)
    locked_rows = []
    unlocked_rows = []

    for _, row in asset_securities.iterrows():
        sec = row["Security Name"]
        key = f"{asset}_{sec}"
        current = row["Market Value (CAD)"]
        method = sec_methods.get(key, "%")
        val = sec_inputs.get(key, 0)
        if sec_locks.get(key, False):
            if method == "%":
                target = val / 100 * asset_target
            elif method == "$":
                target = val
            elif method == "$ Δ":
                target = current + val
            locked_rows.append({**row, "Target $": target})
        else:
            unlocked_rows.append(row)

    locked_df = pd.DataFrame(locked_rows)
    unlocked_df = pd.DataFrame(unlocked_rows)
    locked_total = locked_df["Target $"] .sum() if not locked_df.empty else 0

    if not unlocked_df.empty:
        subtotal = unlocked_df["Market Value (CAD)"].sum()
        unlocked_df["Target $"] = (unlocked_df["Market Value (CAD)"] / subtotal) * (asset_target - locked_total)

    final = pd.concat([locked_df, unlocked_df])
    final["Buy/Sell $"] = final["Target $"] - final["Market Value (CAD)"]
    results.append(final)

security_result_df = pd.concat(results)

# === Display results ===
# === Target Allocation Warning ===
if grouped["Target %"].sum() > 100.5:
    st.warning("⚠️ Total target allocation exceeds 100%. Please adjust your inputs.")
st.subheader("📥 Asset Class Rebalancing Plan")
asset_display_df = grouped.set_index("Asset Class").reindex(desired_order).dropna(how='all').reset_index()
asset_display_df["Current $"] = asset_display_df["Current $"].apply(lambda x: f"${x:,.2f}")
asset_display_df["Target $"] = asset_display_df["Target $"].apply(lambda x: f"${x:,.2f}")
asset_display_df["Buy/Sell $"] = asset_display_df["Buy/Sell $"].apply(lambda x: f"${x:,.2f}")
asset_display_df["Current %"] = asset_display_df["Current %"].apply(lambda x: f"{x:.2f}%")
asset_display_df["Target %"] = asset_display_df["Target %"].apply(lambda x: f"{x:.2f}%")
st.dataframe(asset_display_df.set_index("Asset Class"), use_container_width=True)

# === Security-Level Rebalancing Plan ===
st.subheader("📌 Security-Level Rebalancing Plan")
for asset in desired_order:
    sec_subset = security_result_df[security_result_df["Asset Class"] == asset]
    if not sec_subset.empty:
        with st.expander(f"🔽 Securities in {asset}"):
            display_securities = sec_subset.copy()
            display_securities["Current % of Class"] = (display_securities["Market Value (CAD)"] / display_securities["Market Value (CAD)"].sum()) * 100
            display_securities["Target % of Class"] = (display_securities["Target $"] / display_securities["Target $"] .sum()) * 100
            display_securities = display_securities.round(2)
            display_securities_display = display_securities[[
                "Security Name",
                "Market Value (CAD)",
                "Current % of Class",
                "Target $",
                "Target % of Class",
                "Buy/Sell $"
            ]]
            st.dataframe(display_securities_display.set_index("Security Name"), use_container_width=True)

# === Allocation Pie Charts ===
st.subheader("📊 Allocation Charts")

col1, col2 = st.columns(2)

with col1:
    st.markdown("<h5 style='text-align:center;'>Current Allocation</h5>", unsafe_allow_html=True)
    current_values = grouped.set_index("Asset Class")["Current $"]
    plt.figure(figsize=(4.5, 4.5))
    plt.pie(current_values, labels=current_values.index, autopct="%1.1f%%", startangle=90, colors=["#c9b9a6", "#bfb7ae", "#a69e8c", "#d8d2ca"])
    plt.axis("equal")
    st.pyplot(plt.gcf())

with col2:
    st.markdown("<h5 style='text-align:center;'>Target Allocation</h5>", unsafe_allow_html=True)
    target_values = grouped.set_index("Asset Class")["Target $"]
    plt.figure(figsize=(4.5, 4.5))
    plt.pie(target_values, labels=target_values.index, autopct="%1.1f%%", startangle=90, colors=["#c9b9a6", "#bfb7ae", "#a69e8c", "#d8d2ca"])
    plt.axis("equal")
    st.pyplot(plt.gcf())

# === Summary Section ===
st.subheader("📝 Summary")

buys = grouped[grouped["Buy/Sell $"] > 0]
sells = grouped[grouped["Buy/Sell $"] < 0]
col1, col2 = st.columns(2)
with col1:
    st.markdown("### 🟢 Buy (Asset Classes)")
    for _, row in buys.iterrows():
        st.markdown(f"• **{row['Asset Class']}**: ${row['Buy/Sell $']:,.2f}")
with col2:
    st.markdown("### 🔴 Sell (Asset Classes)")
    for _, row in sells.iterrows():
        st.markdown(f"• **{row['Asset Class']}**: ${abs(row['Buy/Sell $']):,.2f}")

sec_buys = security_result_df[security_result_df["Buy/Sell $"] > 0]
sec_sells = security_result_df[security_result_df["Buy/Sell $"] < 0]
col1, col2 = st.columns(2)
with col1:
    st.markdown("### 🟢 Buy (Securities)")
    for _, row in sec_buys.iterrows():
        st.markdown(f"• **{row['Security Name']} ({row['Asset Class']})**: ${row['Buy/Sell $']:,.2f}")
with col2:
    st.markdown("### 🔴 Sell (Securities)")
    for _, row in sec_sells.iterrows():
        st.markdown(f"• **{row['Security Name']} ({row['Asset Class']})**: ${abs(row['Buy/Sell $']):,.2f}")

# === Download File ===
with pd.ExcelWriter("rebalancing_plan.xlsx") as writer:
    grouped.to_excel(writer, sheet_name="Asset Class", index=False)
    security_result_df.to_excel(writer, sheet_name="Securities", index=False)
with open("rebalancing_plan.xlsx", "rb") as f:
    st.download_button("⬇️ Download Excel", f, file_name="rebalancing_plan.xlsx")
